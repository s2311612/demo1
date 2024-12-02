import openpyxl
from copy import copy
import openpyxl.cell
# 定义一个函数来复制单元格数据和样式
def copy_cell(source_cell, target_cell):
    if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
        target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


# 复制合并单元格并处理其数据、样式和边框
def copy_merged_cells(source_sheet, target_sheet):
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))
        # 获取合并区域左上角单元格
        min_row, min_col, max_row, max_col = merged_cell_range.bounds
        source_top_left_cell = source_sheet.cell(row=min_row, column=min_col)
        target_top_left_cell = target_sheet.cell(row=min_row, column=min_col)
        # 复制左上角单元格的数据和样式
        copy_cell(source_top_left_cell, target_top_left_cell)

        # 复制合并单元格的所有单元格的边框
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                source_cell = source_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(row=row, column=col)
                if source_cell.has_style:
                    target_cell.border = copy(source_cell.border)