import openpyxl
import datetime
from copy import copy
from openpyxl import cell
from copyExcelStyle import copy_cell, copy_merged_cells

# 打开A表
workbook_A = openpyxl.load_workbook(r'C:\Users\郑大爷的小本子\Desktop\国显2024年11月份考勤.xlsx')
s1 = workbook_A['总表']
work_date_l = 'H3'
work_model_l = 'B7'
work_quantity_l = 'F7'
defects_quantity_l = 'H7'
total_hours_l = 'F39'
gaffer_l = 'C5'
workers_l = 'G5'
work_end_style_l = 'J7'
work_hours_l = 'F14'
work_times_l = 'J4'
work_locations_l = 'J3'
work_content_l = 'I7'


# 通用的读取数据函数
def read_data(sheet, min_col, max_col, min_row=5, max_row=47, transform=None):
    data = []
    for row in sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row):
        for cell in row:
            if transform:
                data.append(transform(cell.value))
            else:
                data.append(cell.value)
    return data


# 日期转换函数
def transform_date(excel_date):
    date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
    return date.strftime('%m-%d')


# 日期转换函数
def transform_dates(excel_date):
    date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
    return date.strftime('%Y/%m/%d')


# def transform_dates(excel_date):
#     if excel_date is not None and isinstance(excel_date, (int,float)):  # 确保日期是数值类型且不为None
#         date = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
#         return date.strftime('%Y/%m/%d')
#     return None

# 读取总表的日期并储存
e_dates = read_data(s1, min_col=1, max_col=1, transform=transform_date)

# 单独储存日期给表值
work_dates = read_data(s1, min_col=1, max_col=1, transform=transform_dates)

# 读取总表的作业地点并储存
work_locations = read_data(s1, min_col=10, max_col=10)

# 读取总表的作业内容并储存
work_contents = read_data(s1, min_col=4, max_col=4)

# 读取总表型号并储存
work_models = read_data(s1, min_col=2, max_col=2)

# 读取班别信息
work_times = read_data(s1, min_col=3, max_col=3)

# 读取作业数量并储存
work_quantity = read_data(s1, min_col=5, max_col=5)

# 读取不良数量并储存
defects_quantity = read_data(s1, min_col=6, max_col=6)

# 工时数
work_hours = read_data(s1, min_col=8, max_col=8)

# 工作人数
workers = read_data(s1, min_col=9, max_col=9)

# 领班
gaffer = read_data(s1, min_col=11, max_col=11)

# 总工时
total_hours = read_data(s1, min_col=12, max_col=12)

# 作业结算方式
work_end_style = read_data(s1, min_col=13, max_col=13)

# 获取现有工作表数量
existing_sheets = workbook_A.sheetnames

# 源数据表，后面都是复制此表全部数据后修改
source_sheet = workbook_A[existing_sheets[1]]

# 如果工作表少于总表日期则新建
for i in range(len(existing_sheets), len(e_dates) + 1):
    workbook_A.create_sheet(title=f'new_{i + 1}')

# 遍历从第二张表开始的所有工作表并复制数据、样式和合并单元格
for sheet_name in existing_sheets[2:]:
    target_sheet = workbook_A[sheet_name]
    target_sheet.delete_rows(1, target_sheet.max_row)  # 清空目标表中的现有数据
    copy_merged_cells(source_sheet, target_sheet)  # 复制合并单元格
    for row in source_sheet.iter_rows():
        for source_cell in row:
            if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):  # 确保不是MergedCell
                target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column)
                copy_cell(source_cell, target_cell)

# 遍历所有工作表并修改名称
combined_names = [f'{w_date}{w_work_times}{w_location}{w_content}' for w_date,w_work_times, w_location, w_content in
                  zip(e_dates,work_times, work_locations, work_contents)]
for index, sheet_name in enumerate(workbook_A.sheetnames[1:], start=2):
    try:
        sheet = workbook_A[sheet_name]
        index = index - 2  # 索引位置变动了，需要手动重置
        sheet_name = combined_names[index]
        sheet.title = sheet_name
        sheet[work_date_l].value = work_dates[index]  # 修改格子中的日期
        sheet[work_model_l].value = work_models[index]  # 修改型号
        sheet[work_times_l].value = work_times[index]  # 修改工作班别
        sheet[work_quantity_l].value = work_quantity[index]  # 修改作业数量
        sheet[defects_quantity_l].value = defects_quantity[index]  # 修改不良数量
        sheet[total_hours_l].value = total_hours[index]  # 修改总工时
        sheet[gaffer_l].value = gaffer[index]  # 修改领班
        sheet['B14'].value = gaffer[index]  # 修改的第一个员工名
        sheet[workers_l].value = workers[index]  # 修改工作人数
        sheet['J5'].value = workers[index]  # 修改机器数，用的工作人数
        sheet[work_end_style_l].value = work_end_style[index]  # 修改结算方式
        sheet['G14'].value = work_end_style[index]  # 修改结算方式
        sheet[work_hours_l].value = work_hours[index]  # 修改工时数
        sheet[work_locations_l].value = work_locations[index]  # 修改工作地址
        sheet[work_content_l].value = work_contents[index]  # 工作内容填充
        sheet['C14'].value = work_contents[index]  # 工作内容填充
    except IndexError:
        print(f'超出索引范围{index}')

workbook_A.save(r'C:\Users\郑大爷的小本子\Desktop\国显2024年11月份考勤.xlsx')
print('已操作完毕')
